"""
This file aims to
generate origin and RTStructures mha(s) of all patients.

Author: Han
"""
import multiprocessing
import time

import numpy as np
from openpyxl import load_workbook
from loguru import logger

import preConfig as pc
from fileProcess import get_min_max_layers_from_excel
from generator import Generator


def sub_process_generate_mha(path_list, lock_path_list, lock_excel, total, p_name, logger_name):
    """
    A process to generate mha.
    :param path_list: the paths of the folders of all patients
    :param lock_path_list: process lock locking "path_list"
    :param lock_excel: process lock locking excel
    :param total: total length of original "path_list"
    :param p_name: process number
    :param logger_name: logger to be written to
    :return: None
    """
    logger.add(logger_name, enqueue=True)
    while True:
        lock_path_list.acquire()
        if not path_list:
            lock_path_list.release()
            break
        input_path_g = path_list.pop()
        lock_path_list.release()

        package_name = input_path_g.split('/')[-1]
        patient_id = int(package_name.split('-')[1])
        length = len(path_list)

        tmp_start = time.time()
        logger.info(f'[{p_name}] {round((total - length) * 100 / total):>3}% Processing: {package_name}')

        try:
            generator = Generator(input_path_g, pc.output_path)

            """Generate the mha(s)."""
            # rtStructure = generator.get_rs()
            # generator.generate_rs_and_origin_mha(rtStructure)
            # rtStructure = generator.get_rs_from_mha(pc.output_path, pc.input_type)
            # generator.generate_mha(rtStructure, 'rs_cropped')
            # generator.generate_origin_mha()

            """Write the number of min, max and total layers into the excel."""
            # min_layer, max_layer = generator.get_min_max_layer(rtStructure)
            # total_layer = generator.origin.length
            # lock_excel.acquire()
            # write_layers_to_excel(excel_path, patient_id, min_layer, max_layer, total_layer)
            # lock_excel.release()

            """Crop the mha(s)."""
            # image3D = generator.get_mha(pc.output_path, pc.input_type)
            # image3D = generator.get_mha(pc.output_path)
            # min_layer, max_layer = get_min_max_layers_from_excel(pc.excel_path, patient_id)
            # total_layer = generator.origin.length
            # roi_square = pc.roi_square
            # roi_square[0] = [min_layer, max_layer]
            # image3D = np.delete(image3D, [z for z in range(0, roi_square[0][0]-1)] +
            #                     [z for z in range(roi_square[0][1], total_layer)], axis=0)
            # image3D = np.delete(image3D, [y for y in range(0, roi_square[1][1]-1)] +
            #                     [y for y in range(roi_square[2][1], 512)], axis=1)
            # image3D = np.delete(image3D, [x for x in range(0, roi_square[1][0]-1)] +
            #                     [x for x in range(roi_square[2][0], 512)], axis=2)
            # generator.generate_mha(image3D, 'o_cropped')

        except Exception as e:
            logger.error(f'[{p_name}] {package_name}: {e}')
        logger.info(f"--[{p_name}] {package_name} cost: {str(time.time() - tmp_start).split('.')[0]}s")
    logger.info(f'----<[{p_name}] finished.>')


def main_process():
    """
    The main process controling the sub-processes.
    :return: None
    """
    manager = multiprocessing.Manager()
    path_list = manager.list()
    lock_path_list = manager.Lock()
    lock_excel = manager.Lock()

    """Get all the patients' paths."""
    excel_clinical = pc.excel_path
    wb = load_workbook(excel_clinical)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        path_list.append(pc.input_path + '/' + row[2].value + '-'
                         + f'{row[1].value:010}')

    """Start all sub processes."""
    total = len(path_list)
    processes = []
    logger_name = f'log/log_{time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())}.log'
    for i in range(pc.process_num):
        p_name = str(i)
        p = multiprocessing.Process(target=sub_process_generate_mha,
                                    args=(path_list, lock_path_list, lock_excel, total, p_name, logger_name))
        p.start()
        processes.append(p)
    for p in processes:
        p.join()
    logger.add(logger_name, enqueue=True)
    logger.info('<Main process end.>')


if __name__ == '__main__':
    main_process()
