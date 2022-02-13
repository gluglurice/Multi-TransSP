"""
This file aims to
regularize the names of folders of dicoms.

Authors: Han
"""
import os
import pydicom
import pandas as pd
from openpyxl import load_workbook
from loguru import logger

import preConfig as pc


def modify_folder_names():
    """
    Rename the folders
    :return: None
    """
    for root, dirs, files in os.walk(pc.input_path):
        for patient in dirs:
            try:
                file = os.path.join(root, patient, os.listdir(os.path.join(root, patient))[0])
                dcm = pydicom.read_file(file)
                patient_id = str(dcm.PatientID)
                patient_name = str(dcm.PatientName)
                os.rename(os.path.join(root, patient), os.path.join(root, patient_name + '-' + patient_id))
            except Exception as e:
                logger.error(f'{patient}: {e}')


def get_min_max_layers_from_excel(excel_clinical, patient_id):
    """
    Get the number tuple of min and max layers from the excel according to patient_id.
    :return:
        min_layer: number of the min layer which contains rs info.
        max_layer: number of the max layer which contains rs info.
    """
    wb = load_workbook(excel_clinical)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[pc.col_id - 1].value == patient_id:
            min_layer = row[pc.col_min_layer - 1].value
            max_layer = row[pc.col_max_layer - 1].value
            return min_layer, max_layer


def write_layers_to_excel(excel_clinical, patient_id, min_layer, max_layer, total_layer):
    """
    Write the number of min and max layers into the excel.
    :return: None
    """
    wb = load_workbook(excel_clinical)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[pc.col_id - 1].value == patient_id:
            row[pc.col_min_layer - 1].value = min_layer
            row[pc.col_max_layer - 1].value = max_layer
            row[pc.col_total_layer - 1].value = total_layer
            break
    wb.save(pc.excel_path)


def fill_excel_stage(excel_clinical):
    """
    Map the stage info into classification numbers.
    :return: None
    """
    wb = load_workbook(excel_clinical)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        T = row[pc.col_stage_T - 1].value
        N = row[pc.col_stage_N - 1].value
        if T is None or N is None:
            continue
        elif T == 'T1' and N == 'N0':
            row[pc.col_stage - 1].value = 1
        elif (T == 'T2' and N == 'N0') or ((T == 'T1' or T == 'T2') and N == 'N1'):
            row[pc.col_stage - 1].value = 2
        elif ((T == 'T1' or T == 'T2') and N == 'N2') or (T == 'T3' and (N == 'N0' or N == 'N1' or N == 'N2')):
            row[pc.col_stage - 1].value = 3
        elif T == 'T4' and (N == 'N0' or N == 'N1' or N == 'N2'):
            row[pc.col_stage - 1].value = 4
        elif N.startswith('N3'):
            row[pc.col_stage - 1].value = 5
    wb.save(excel_clinical)


def fill_excel_missing_value(excel_clinical):
    """
    Fill the missing value with mean or mode.
    :return: None
    """
    wb = load_workbook(excel_clinical)
    ws = wb[wb.sheetnames[0]]
    rt_days_list = []
    BMI_list = []
    stage_list = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        rt_days_list.append(row[pc.col_rt_days - 1].value)
        BMI_list.append(row[pc.col_BMI - 1].value)
        stage_list.append(row[pc.col_stage - 1].value)
    data_mean = {'rt_days': rt_days_list, 'BMI': BMI_list, 'stage': stage_list}
    data_mean_frame = pd.DataFrame(data_mean)
    data_mean_frame = data_mean_frame.fillna(data_mean_frame.mean())
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        if row[pc.col_rt_days - 1].value is None:
            row[pc.col_rt_days - 1].value = data_mean_frame['rt_days'][i]
        if row[pc.col_BMI - 1].value is None:
            row[pc.col_BMI - 1].value = data_mean_frame['BMI'][i]
        if row[pc.col_stage - 1].value is None:
            row[pc.col_stage - 1].value = 3
    wb.save(excel_clinical)


def main():
    modify_folder_names()
    # fill_excel_stage(preprocess_config.excel_path)
    # fill_excel_missing_value(preprocess_config.excel_path)


if __name__ == '__main__':
    main()
