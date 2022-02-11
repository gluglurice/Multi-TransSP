"""
MyDataset class

Authors: Han
"""
import glob
import random

import numpy as np
import torch
from torch.utils.data import Dataset
import SimpleITK as sitk
from openpyxl import load_workbook

import config as c


class MyDataset(Dataset):
    """
    Load radiographs(mha(s) here) and clinical data(excel).
    """

    def __init__(self, root, excel_path, mode='train', ki=1, k=1, transform=None, rand=False):
        """
        Initialize the class.
        :param root: mha data root
        :param excel_path: path of the excel containing text features
        :param mode: "train", "validate", or "test"
        :param ki: the iter number of KFold, ranging from [0, k-1]
        :param k: the total number of KFold, usually is 5, but never 0
               Especially, when ki==k==1, "validate" list will be None.
        :param transform: transforms_train means
        :param rand: whether shuffle the dataset (with a constant seed)
        """
        self.root = root
        self.excel_path = excel_path
        self.mode = mode
        self.ki = ki
        self.k = k
        self.transform = transform
        self.rand = rand
        self.mha_list = self.get_mha_list()
        self.row_list = self.get_row_list()
        self.max_valid_slice_num = self.get_max_valid_slice_num()

    def __len__(self):
        return len(self.mha_list)

    def __getitem__(self, index):
        """
        :param index: the index of the patient in mha_list
        :return:
            {
            'image3D': image3D, (image feature, torch.Size([1, z, h, w]))
            'text': text, (text feature, torch.Size([1, len])
            'survivals': survivals (regression labels, torch.Size([1, len])
            }
        """
        mha = self.mha_list[index]
        image3D = sitk.GetArrayFromImage(sitk.ReadImage(mha)).squeeze().astype('float32')
        image3D = (image3D - np.min(image3D)) / (np.max(image3D) - np.min(image3D))
        image3D = torch.from_numpy(image3D)
        if self.transform:
            image3D = self.transform(image3D)

        row = self.row_list[index]

        text_list = []
        for i in range(c.col_feature_start - 1, c.col_feature_end):
            text_list.append(row[i].value)
        text = torch.tensor(text_list, dtype=torch.float)

        survival_list = []
        for i in range(c.col_label_start - 1, c.col_label_end):
            survival_list.append(row[i].value)
        survivals = torch.tensor(survival_list, dtype=torch.float)

        return {'mha': mha, 'image3D': image3D, 'text': text, 'survivals': survivals}

    def get_mha_list(self):
        """
        Get an mha list according to mode. When mode is "train" or "validate",
        this function returns an mha list which is K-folded guided by ki and k.
        And when mode is "test", it returns test mha list.
        :return:
            mha_list: an mha list
        """
        mha_list = glob.glob(self.root + c.mha_files_path)
        if self.rand:
            random.seed(c.seed)
            random.shuffle(mha_list)

        total_len = len(mha_list)
        train_len = total_len - total_len // c.train_test_ratio
        every_fold_len = train_len // self.k

        mha_train_list = mha_list[:train_len]
        mha_test_list = mha_list[train_len:]

        if self.mode == 'train':
            mha_list = mha_train_list[:every_fold_len * self.ki] + \
                       mha_train_list[every_fold_len * (self.ki + 1):]
        elif self.mode == 'validate':
            mha_list = mha_train_list[every_fold_len * self.ki:
                                      every_fold_len * (self.ki + 1)]
        elif self.mode == 'test':
            mha_list = mha_test_list

        return mha_list

    def get_row_list(self):
        """
        :return:
            row_list: the excel row list of each patient, sorted according to mha_list
        """
        row_list = []
        wb = load_workbook(self.excel_path)
        ws = wb[wb.sheetnames[0]]
        for col in ws.iter_cols(min_col=c.col_feature_start, max_col=c.col_label_end):
            col_data = []
            for i in range(1, ws.max_row):
                col_data.append(col[i].value)
            min_value = min(col_data)
            max_value = max(col_data)
            for i in range(1, ws.max_row):
                col[i].value = (col[i].value - min_value) / (max_value - min_value)
        for mha in self.mha_list:
            patient_id = int(mha.split('-')[1].split('.')[0])
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if patient_id == row[c.col_id - 1].value:
                    row_list.append(row)
                    break
        return row_list

    def get_max_valid_slice_num(self):
        """
        :return:
            max_valid_slice_num: the max total layer of all the patients
        """
        max_valid_slice_num = 0
        for row in self.row_list:
            if row[c.col_total_layer - 1].value > max_valid_slice_num:
                max_valid_slice_num = row[c.col_total_layer - 1].value
        return max_valid_slice_num
